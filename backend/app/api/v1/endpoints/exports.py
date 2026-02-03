from io import BytesIO, StringIO
from datetime import datetime
import csv

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.v1.endpoints._deps import get_db, require_admin
from app.models.appointment import Appointment

router = APIRouter()

def _safe(s):
    return "" if s is None else str(s)

@router.get("/admin/appointments/export")
def export_appointments(
    format: str = Query(default="xlsx", pattern="^(xlsx|pdf|csv)$"),
    db: Session = Depends(get_db),
    _admin=Depends(require_admin),
):
    items = db.query(Appointment).order_by(Appointment.id.desc()).all()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"appointments_{ts}"

    if format == "csv":
        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow(["id", "counseling_type", "name", "phone", "address", "message", "status", "created_at"])
        for a in items:
            writer.writerow([
                a.id,
                _safe(getattr(a, "counseling_type", "")),
                _safe(a.name),
                _safe(a.phone),
                _safe(a.address),
                _safe(a.message),
                _safe(a.status),
                _safe(getattr(a, "created_at", "")),
            ])
        bio = BytesIO(sio.getvalue().encode("utf-8"))
        bio.seek(0)
        return StreamingResponse(
            bio,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{base_name}.csv"'}
        )

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Appointments"

        headers = ["ID", "Counseling Type", "Name", "Phone", "Address", "Message", "Status", "Created At"]
        ws.append(headers)

        for a in items:
            ws.append([
                a.id,
                _safe(getattr(a, "counseling_type", "")),
                _safe(a.name),
                _safe(a.phone),
                _safe(a.address),
                _safe(a.message),
                _safe(a.status),
                _safe(getattr(a, "created_at", "")),
            ])

        # Basic autosize (rough)
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 18

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{base_name}.xlsx"'}
        )

    # pdf
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas

    out = BytesIO()
    c = canvas.Canvas(out, pagesize=landscape(A4))
    width, height = landscape(A4)

    c.setFont("Helvetica-Bold", 14)
    c.drawString(30, height - 30, "Kanglei Career Solution - Appointments Export")
    c.setFont("Helvetica", 10)
    c.drawString(30, height - 50, f"Generated: {datetime.now().isoformat(sep=' ', timespec='seconds')}")

    # Table-ish rendering (simple, readable)
    y = height - 80
    c.setFont("Helvetica-Bold", 9)
    cols = ["ID", "Type", "Name", "Phone", "Status", "Created At"]
    x_positions = [30, 80, 210, 360, 470, 560]
    for i, col in enumerate(cols):
        c.drawString(x_positions[i], y, col)

    c.setFont("Helvetica", 9)
    y -= 16

    for a in items[:500]:  # safety cap for PDF
        if y < 40:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica-Bold", 9)
            for i, col in enumerate(cols):
                c.drawString(x_positions[i], y, col)
            c.setFont("Helvetica", 9)
            y -= 16

        c.drawString(x_positions[0], y, str(a.id))
        c.drawString(x_positions[1], y, _safe(getattr(a, "counseling_type", ""))[:18])
        c.drawString(x_positions[2], y, _safe(a.name)[:22])
        c.drawString(x_positions[3], y, _safe(a.phone)[:16])
        c.drawString(x_positions[4], y, _safe(a.status)[:10])
        c.drawString(x_positions[5], y, _safe(getattr(a, "created_at", ""))[:19])
        y -= 14

    c.showPage()
    c.save()
    out.seek(0)

    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{base_name}.pdf"'}
    )
